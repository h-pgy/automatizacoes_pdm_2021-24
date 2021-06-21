from openpyxl import load_workbook
import os
from functools import partial
import pandas as pd
import json
import datetime


class AbstractParserFichas:

    def get_files(self, path_dados=None):

        if path_dados is None:
            path_dados = self.path_dados

        files = [os.path.join(path_dados, f) for f
                 in os.listdir(path_dados) if 'xls' in f.split('.')[-1]]

        return files

    def wb_generator(self, path_dados):

        files = self.get_files(path_dados)

        for file in files:
            yield file, load_workbook(file, read_only=True, data_only=True)

    def get_sheet_by_name(self, wb, sheet_name_padrao):

        for sheet_name in wb.sheetnames:
            if sheet_name_padrao in sheet_name.lower():
                return wb[sheet_name]
        else:
            raise RuntimeError(f'Sheetname {sheet_name_padrao} nao encontrado na planilha!')

    def get_cell_value(self, cell):
        #aqui que eu vou mudar depois pra fazer type enforcement

        valor = cell.value
        if valor is None:
            return ''
        if type(valor) is datetime.datetime:
            return str(valor)

        return valor


    def extract_with_col_mapper(self, wb, mapper):

        sheet = self.get_sheet_by_name(wb, mapper['sheet_name'])

        data = []

        for i in range(mapper['line_ini'], mapper['line_end'] + 1):
            parsed = {}
            for name, col in mapper['data_cells'].items():
                cell = col + str(i)
                parsed[name] = self.get_cell_value(sheet[cell])
            data.append(parsed)

        return data

    def extract_with_col_mapper_variable_rows(self, wb, mapper):

        sheet = self.get_sheet_by_name(wb, mapper['sheet_name'])

        data = []
        i = mapper['line_ini']
        while True:
            parsed = {}
            for name, col in mapper['data_cells'].items():
                cell = col + str(i)
                parsed[name] = self.get_cell_value(sheet[cell])
            check_stop = parsed[mapper['col_stop']]
            if check_stop == '' or check_stop is None:
                break
            data.append(parsed)
            i += 1

        return data

    def extract_with_cell_mapper(self, wb, mapper):

        sheet = self.get_sheet_by_name(wb, mapper['sheet_name'])

        parsed = {}

        for name, cell in mapper['data_cells'].items():
            parsed[name] = self.get_cell_value(sheet[cell])

        return parsed

    def extract_whole_sheet(self, file, sheet_name):

        return pd.read_excel(file, sheet_name=sheet_name,
                             parse_dates=False, thousands=',').to_dict(orient='records')

    def salvar_file(self, parsed_ficha, path_salvar=None):

        f_name = os.path.split(parsed_ficha['file_original'])[-1].replace('.xlsx', '.json')

        if path_salvar is None:
            path_salvar = self.path_salvar

        if not os.path.exists(path_salvar):
            os.mkdir(path_salvar)

        path_file = os.path.join(path_salvar, f_name)
        try:
            with open(path_file, 'w', encoding='cp1252') as f:
                json.dump(parsed_ficha, f, ensure_ascii=False)
        except UnicodeEncodeError as e:
            with open(path_file, 'w', encoding='utf-8') as f:
                json.dump(parsed_ficha, f, ensure_ascii=False)
                print(f'Erro Unicode na file {f_name}')
                print(e)


class ParserFichas(AbstractParserFichas):
    mapper_ficha_tecnica = dict(
        sheet_name='ficha técnica meta',
        data_cells=dict(
            secretaria='c4',
            numero_meta='c5',
            desc_meta='c6',
            indicador='c9',
            contexto='c10',
            info_compl='c11',
            valor_base='c12',
            serie_historica='c13',
            frequencia='c14',
            periodo='c15',
            forma_apuracao='c18',
            observacoes='c20',
            prev_21='c23',
            prev_22='c24',
            prev_23='c25',
            prev_24='c26',
            impacto_covid='b29'
        )
    )

    mapper_iniciativas = dict(
        sheet_name='iniciativas',
        line_ini=4,
        col_stop='descricao',
        data_cells={
            'id': 'b',
            'descricao': 'c',
            'orgao_unidade': 'd',
            'detalhamento': 'e',
            'valor_global': 'f'
        }
    )

    mapper_alteracoes = dict(
        sheet_name='proposta de alteração',
        line_ini=4,
        col_stop='meta_ou_iniciativa',
        data_cells={
            'meta_ou_iniciativa': 'b',
            'item': 'c',
            'versao_publicada': 'd',
            'proposta_alteracao': 'e',
            'justificativa': 'f'
        }
    )

    mapper_regionalizacao = dict(
        sheet_name='regionalização',
        line_ini=3,
        col_stop='subprefeitura', #precisa puxar tudo
        data_cells={
            'subprefeitura': 'b',
            'projecao_quadrienio': 'c',
            'status_regionalizacao': 'd'
        }
    )

    def __init__(self, path_dados, path_salvar):

        self.path_dados = path_dados
        self.files = self.get_files(path_dados)
        self.path_salvar = path_salvar
        self.set_methods()

    def set_methods(self):

        self.extract_ficha_tecnica = partial(self.extract_with_cell_mapper,
                                             mapper=self.mapper_ficha_tecnica)
        self.extract_iniciativas = partial(self.extract_with_col_mapper_variable_rows,
                                           mapper=self.mapper_iniciativas)
        self.extract_alteracoes = partial(self.extract_with_col_mapper_variable_rows,
                                          mapper=self.mapper_alteracoes)
        self.extract_regionalizacao = partial(self.extract_with_col_mapper_variable_rows,
                                              mapper=self.mapper_regionalizacao)
        self.extract_custos = partial(self.extract_whole_sheet, sheet_name='Custo')
        self.extract_orcamento = partial(self.extract_whole_sheet, sheet_name='Orçamento')

    def extract_ficha(self, file, wb=None, salvar=False, path_salvar=None):

        if wb is None:
            wb = load_workbook(file)

        parsed = {
            'ficha_tecnica': self.extract_ficha_tecnica(wb),
            'iniciativas': self.extract_iniciativas(wb),
            'alteracoes': self.extract_alteracoes(wb),
            'regionalizacao': self.extract_regionalizacao(wb),
            'custos': self.extract_custos(file),
            'orcamento': self.extract_orcamento(file),
            'file_original': file
        }

        if salvar:
            self.salvar_file(parsed, path_salvar)

        return parsed

    def extract_all_fichas(self, path_dados=None, salvar=True, path_salvar=None):

        if path_dados is None:
            path_dados = self.path_dados
        parsed_data = []
        wbs = self.wb_generator(path_dados)
        for file, wb in wbs:
            print(file)
            ficha = self.extract_ficha(file, wb, salvar, path_salvar)
            parsed_data.append(ficha)

        return parsed_data

    def __call__(self):

        return self.extract_all_fichas()